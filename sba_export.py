#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Aug 29 19:12:53 2025

@author: harshadghodke
"""

# sba_export.py
# Drop this alongside your other modules (e.g., next to modular_simulator.py)

from __future__ import annotations
import os, json, hashlib, datetime, math
from typing import Dict, Any, List, Optional
from openpyxl import load_workbook

# ---------- Minimal data contract ----------
# Expect at least:
# df_monthly: pandas.DataFrame with monthly rows (0..MONTHS-1) and columns like:
#   revenue_total, cogs_total, opex_total, interest_expense, depreciation_expense,
#   cash_begin, cash_end, cash_inflows, cash_outflows,
#   loan_balance_504, loan_balance_7a, loan_payment_504, loan_payment_7a,
#   capex_purchases, loan_draw_504, loan_draw_7a, ...
# capex_items: list of dicts with fields: {"label", "amount", "category"} where category ∈
#   {"leasehold", "equipment", "furniture", "vehicles", "other"}
# loans: dict with {"504": {"principal_used", "rate", "term_years", "io_months"},
#                   "7a": {"principal_used", "rate", "term_years", "io_months"}}
# payroll_cfg: {"owner_salary_monthly": float, "payroll_tax_rate": float,
#               "staff": [{"role": str, "hours_per_week": float, "hourly_rate": float,
#                          "start_month": int}]}
# working_capital: float (target buffer / OPEX loan size)
# equity_injection: float
# lease_years: int (for depreciation of leasehold improvements)
# dep_life_equipment_years: int
# dep_life_furniture_years: int
# NOTE: All fields are optional; writer gracefully skips missing pieces.

import pandas as pd

def _ws(wb, name: str):
    """Safe sheet accessor: returns worksheet or None."""
    return wb[name] if name in wb.sheetnames else None

def _short_hash(obj: Any) -> str:
    try:
        s = json.dumps(obj, sort_keys=True, default=str)
    except Exception:
        s = str(obj)
    return hashlib.sha1(s.encode("utf-8")).hexdigest()[:8]

def _now_ts() -> str:
    return datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

def make_run_id(config_like: Dict[str, Any], seed_like: Any = None) -> str:
    base = {
        "ts": _now_ts(),
        "h": _short_hash({"cfg": config_like, "seed": seed_like}),
    }
    return f'{base["ts"]}_{base["h"]}'

def _safe_set(ws, cell: str, value):
    try:
        ws[cell] = value
    except Exception:
        pass

def _sum_or_zero(vals: List[float]) -> float:
    return float(sum(v for v in vals if v is not None))

def _annual_sum(df, col: str, year_index: int) -> float:
    # year_index: 0-> months 0-11; 1->12-23; 2->24-35
    start = year_index * 12
    end = start + 12
    if col not in df.columns: return 0.0
    sub = df[col].iloc[start:end] if len(df) >= end else df[col].iloc[start:len(df)]
    return float(sub.fillna(0).sum())

def _year_end(df, col: str, year_index: int) -> float:
    idx = min((year_index + 1) * 12 - 1, len(df) - 1)
    if col not in df.columns or idx < 0: return 0.0
    v = df[col].iloc[idx]
    return 0.0 if (v is None or (isinstance(v, float) and math.isnan(v))) else float(v)

def _group_capex(capex_items: List[Dict[str, Any]]) -> Dict[str, float]:
    buckets = {"leasehold":0.0, "equipment":0.0, "furniture":0.0, "vehicles":0.0, "other":0.0}
    for it in capex_items or []:
        cat = (it.get("category") or "other").lower()
        amt = float(it.get("amount") or 0.0)
        if cat not in buckets: cat = "other"
        buckets[cat] += amt
    return buckets

def _build_depreciation(capex_buckets: Dict[str, float],
                        lease_years: int,
                        dep_life_equipment_years: int,
                        dep_life_furniture_years: int) -> Dict[str, float]:
    """
    Straight-line annual depreciation by major class for P&L; BS uses accumulated via years.
    """
    leasehold = capex_buckets.get("leasehold", 0.0)
    equipment = capex_buckets.get("equipment", 0.0)
    furniture = capex_buckets.get("furniture", 0.0)
    # Vehicles + other -> fold into equipment life for simplicity (conservative)
    vehicles  = capex_buckets.get("vehicles", 0.0)
    other     = capex_buckets.get("other", 0.0)

    dep = {
        "leasehold_annual": (leasehold / max(1, lease_years)) if leasehold else 0.0,
        "equipment_annual": ((equipment + vehicles + other) / max(1, dep_life_equipment_years)) if (equipment+vehicles+other) else 0.0,
        "furniture_annual": (furniture / max(1, dep_life_furniture_years)) if furniture else 0.0,
    }
    dep["total_annual"] = dep["leasehold_annual"] + dep["equipment_annual"] + dep["furniture_annual"]
    dep["total_monthly"] = dep["total_annual"] / 12.0
    return dep

def export_to_sba_workbook(
    df_monthly,                         # pandas.DataFrame (36+ rows recommended)
    template_path: str,                 # "/mnt/data/Financial Projections Template (2).xlsx"
    output_root: str,                   # e.g. "./runs" or "/mnt/data/exports"
    config_snapshot: Dict[str, Any],    # dict of all key inputs/sliders/terms
    capex_items: Optional[List[Dict[str, Any]]] = None,
    loans: Optional[Dict[str, Any]] = None,
    payroll_cfg: Optional[Dict[str, Any]] = None,
    working_capital: Optional[float] = None,
    equity_injection: Optional[float] = None,
    lease_years: int = 10,
    dep_life_equipment_years: int = 5,
    dep_life_furniture_years: int = 7,
    run_seed: Optional[Any] = None,
) -> Dict[str, Any]:
    """
    Creates a per-run folder, copies SBA template with filled values, and writes metadata.json.
    Returns dict with paths and run_id.
    """
    # --- Make run folder -----------------------------------------------------
    run_id = make_run_id(config_snapshot, seed_like=run_seed)
    out_dir = os.path.join(output_root, run_id)
    os.makedirs(out_dir, exist_ok=True)

    # --- Save simulator CSV for reproducibility ------------------------------
    try:
        csv_path = os.path.join(out_dir, "monthly_timeseries.csv")
        df_monthly.to_csv(csv_path, index=False)
    except Exception:
        csv_path = None

    # --- Open template workbook ---------------------------------------------
    wb = load_workbook(template_path)

    # ---------- 1) Startup Costs & Funding ----------
    capex_buckets = _group_capex(capex_items or [])
    equipment_total = capex_buckets.get("equipment", 0.0) + capex_buckets.get("vehicles", 0.0) + capex_buckets.get("other", 0.0)
    leasehold_total = capex_buckets.get("leasehold", 0.0)
    furniture_total = capex_buckets.get("furniture", 0.0)
    total_capex = equipment_total + leasehold_total + furniture_total

    ws = _ws(wb, "1.Startup Costs & Funding") if "1.Startup Costs & Funding" in wb.sheetnames else None
    if ws:
        # NOTE: cell addresses are template-specific; adjust if needed.
        # Use-of-funds (amount column is usually "C" in many SBA templates)
        _safe_set(ws, "C4", leasehold_total)   # Leasehold Improvements
        _safe_set(ws, "C5", furniture_total)   # Furniture/Fixtures
        _safe_set(ws, "C6", equipment_total)   # Equipment
        _safe_set(ws, "C2", total_capex)       # Total Startup Costs (sum)
        # Working capital buffer:
        if working_capital is not None:
            _safe_set(ws, "C9", float(working_capital))  # Working Capital (row varies by template)
        # Sources of funds (loans + equity)
        if loans:
            _safe_set(ws, "C15", float(loans.get("504", {}).get("principal_used", 0.0)))
            _safe_set(ws, "C16", float(loans.get("7a", {}).get("principal_used", 0.0)))
        if equity_injection is not None:
            _safe_set(ws, "C17", float(equity_injection))

    # ---------- 2) Sales Forecast ----------
    ws = _ws(wb, "2.Sales Forecast") if "2.Sales Forecast" in wb.sheetnames else None
    if ws and ("revenue_total" in df_monthly.columns):
        # If you have per-stream columns (e.g., revenue_memberships, revenue_events...), list them here
        streams = [c for c in df_monthly.columns if c.startswith("revenue_") and c != "revenue_total"]
        if not streams:
            # Fallback: a single line “Total Sales”
            # B3 name, F3 total sales (Y1), F4 (Y2), F5 (Y3) etc. — adjust for your exact template layout
            _safe_set(ws, "B3", "Total Sales")
            for yi in range(3):
                _safe_set(ws, f"F{3+yi}", _annual_sum(df_monthly, "revenue_total", yi))
        else:
            row = 3
            for s in streams:
                _safe_set(ws, f"B{row}", s.replace("revenue_", "").replace("_", " ").title())
                for yi in range(3):
                    _safe_set(ws, f"F{row+yi}", _annual_sum(df_monthly, s, yi))
                row += 3  # leave space; adapt to your layout

    # ---------- 3) Payroll ----------
    ws = _ws(wb, "3.Payroll") if "3.Payroll" in wb.sheetnames else None
    if ws and payroll_cfg:
        owner_salary = float(payroll_cfg.get("owner_salary_monthly", 0.0))
        tax_rate = float(payroll_cfg.get("payroll_tax_rate", 0.0))
        # Owner row (example cells)
        _safe_set(ws, "B5", "Owner")
        _safe_set(ws, "C5", owner_salary * 12.0)  # annual
        # Staff (simplified)
        staff = payroll_cfg.get("staff", []) or []
        base_row = 8
        for i, st in enumerate(staff):
            _safe_set(ws, f"B{base_row+i}", st.get("role", f"Staff {i+1}"))
            # compute annualized straight-line cost from start_month onward for Y1..Y3
            # you can refine later; here we write an indicative Year 1 total:
            hrs = float(st.get("hours_per_week", 0.0))
            rate = float(st.get("hourly_rate", 0.0))
            y1_months = max(0, min(12, 12 - int(st.get("start_month", 0))%12))
            _safe_set(ws, f"C{base_row+i}", hrs * rate * 4.333 * y1_months)  # Year 1 estimate
        # Payroll taxes (we’ll show % in assumptions; P&L will include tax expense via df if modeled)

    # ---------- 4) Profit & Loss ----------
    ws = _ws(wb, "4.Profit & Loss") if "4.Profit & Loss" in wb.sheetnames else None
    if ws:
        rev = [ _annual_sum(df_monthly, "revenue_total", yi) for yi in range(3) ]
        cogs = [ _annual_sum(df_monthly, "cogs_total", yi) if "cogs_total" in df_monthly.columns else 0.0 for yi in range(3) ]
        opex = [ _annual_sum(df_monthly, "opex_total", yi) if "opex_total" in df_monthly.columns else 0.0 for yi in range(3) ]
        dep  = [ _annual_sum(df_monthly, "depreciation_expense", yi) if "depreciation_expense" in df_monthly.columns else 0.0 for yi in range(3) ]
        intr = [ _annual_sum(df_monthly, "interest_expense", yi) if "interest_expense" in df_monthly.columns else 0.0 for yi in range(3) ]
        # Example placements (adjust to your template cells)
        for yi in range(3):
            _safe_set(ws, f"C{10+yi}", rev[yi])     # Sales
            _safe_set(ws, f"C{12+yi}", cogs[yi])    # COGS
            _safe_set(ws, f"C{20+yi}", opex[yi])    # Operating expenses (excl. depr & interest if your template separates)
            _safe_set(ws, f"C{22+yi}", dep[yi])     # Depreciation
            _safe_set(ws, f"C{23+yi}", intr[yi])    # Interest

    # ---------- 5) Monthly Cash Flow ----------
    ws = _ws(wb, "5.Monthly Cash Flow") if "5.Monthly Cash Flow" in wb.sheetnames else None
    if ws and len(df_monthly) > 0:
        # Example block: write ending cash for months 1..36 starting at row 25, col C
        for m in range(min(36, len(df_monthly))):
            val = df_monthly.get("cash_end", pd.Series([None]*len(df_monthly))).iloc[m]
            ws.cell(row=25, column=3+m, value=float(val) if val is not None else 0.0)

    # ---------- 6) Balance Sheet ----------
    ws = _ws(wb, "6.Balance Sheet") if "6.Balance Sheet" in wb.sheetnames else None
    if ws:
        # Year-end cash
        for yi in range(3):
            _safe_set(ws, f"C{12+yi}", _year_end(df_monthly, "cash_end", yi))
        # Loans at year-end
        for yi in range(3):
            _safe_set(ws, f"C{20+yi}", _year_end(df_monthly, "loan_balance_504", yi))
            _safe_set(ws, f"C{21+yi}", _year_end(df_monthly, "loan_balance_7a", yi))

    # ---------- 7) Assumptions ----------
    ws = _ws(wb, "7.Assumptions") if "7.Assumptions" in wb.sheetnames else None
    if ws:
        _safe_set(ws, "A2", "Sales")
        _safe_set(ws, "A5", "Operating Expenses")
        _safe_set(ws, "A7", "Fixed Asset Purchases/Capital Expenditures")
        _safe_set(ws, "A9", "Financing Assumptions")
        try:
            _safe_set(ws, "A3", json.dumps({"streams":[c for c in df_monthly.columns if c.startswith("revenue_")]}))
            _safe_set(ws, "A8", json.dumps({"capex_buckets": capex_buckets}))
            _safe_set(ws, "A10", json.dumps({"loans": loans}, default=str))
        except Exception:
            pass

    # --- Save filled workbook ------------------------------------------------
    out_xlsx = os.path.join(out_dir, "Financial_Projections_Filled.xlsx")
    wb.save(out_xlsx)

    # --- Metadata ------------------------------------------------------------
    meta = {
        "run_id": run_id,
        "timestamp": datetime.datetime.now().isoformat(),
        "template_path": template_path,
        "output_root": output_root,
        "out_dir": out_dir,
        "out_xlsx": out_xlsx,
        "csv_path": csv_path,
        "config_snapshot": config_snapshot,
        "loans": loans,
        "working_capital": working_capital,
        "equity_injection": equity_injection,
        "lease_years": lease_years,
        "dep_life_equipment_years": dep_life_equipment_years,
        "dep_life_furniture_years": dep_life_furniture_years,
        "payroll_cfg": payroll_cfg,
    }
    meta_path = os.path.join(out_dir, "metadata.json")
    with open(meta_path, "w") as f:
        json.dump(meta, f, indent=2, default=str)

    return {"run_id": run_id, "out_dir": out_dir, "xlsx": out_xlsx, "metadata": meta_path}